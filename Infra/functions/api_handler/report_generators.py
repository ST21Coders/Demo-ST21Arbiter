"""Report generators — turn a data bundle into downloadable bytes.

generate(report_id, fmt, bundle) -> (payload: bytes, content_type: str, ext: str)

Formats: pdf (reportlab), xlsx (openpyxl), csv/json/zip (stdlib). reportlab and
openpyxl are lazy-imported so the csv/json/zip paths still work even if a heavy
dependency failed to bundle. The bundle shape is built by api_handler._report_bundle:

  {conflicts, audit, change_requests, summaries, overall, org_name,
   generated_at, params}
"""
from __future__ import annotations

import csv
import io
import json
import zipfile
from decimal import Decimal

SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}


class _DecimalEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, Decimal):
            return int(o) if o == o.to_integral_value() else float(o)
        return super().default(o)


def _num(v):
    if isinstance(v, Decimal):
        return int(v) if v == v.to_integral_value() else float(v)
    return v


def _join(v):
    return "; ".join(v) if isinstance(v, list) else (v or "")


def _filter_conflicts(conflicts, params):
    params = params or {}
    sev = {s.upper() for s in (params.get("severity") or [])}
    status = {s.upper() for s in (params.get("status") or [])}
    out = list(conflicts)
    if sev and len(sev) < 4:
        out = [c for c in out if (c.get("severity") or "").upper() in sev]
    if status and len(status) < 3:
        out = [c for c in out if (c.get("status") or "OPEN").upper() in status]
    out.sort(key=lambda c: (SEVERITY_ORDER.get((c.get("severity") or "").upper(), 9),
                            c.get("conflict_id") or ""))
    return out


def _filter_summaries(summaries, params):
    ids = (params or {}).get("frameworks")
    if not ids:
        return summaries
    wanted = set(ids)
    return [s for s in summaries if s["id"] in wanted]


def _top_open_risks(conflicts, n=5):
    open_ = [c for c in conflicts if (c.get("status") or "OPEN").upper() in ("OPEN", "IN_REVIEW")]
    open_.sort(key=lambda c: (SEVERITY_ORDER.get((c.get("severity") or "").upper(), 9),
                              c.get("conflict_id") or ""))
    return open_[:n]


# ──────────────────────────── reportlab (PDF) ────────────────────────────
def _rl():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "PDF generation requires reportlab, which is not available in this "
            "deployment. Use a csv/xlsx/json/zip format, or add reportlab to the "
            f"api_handler bundle. ({e})"
        )
    return locals()


def _styles(rl):
    base = rl["getSampleStyleSheet"]()
    PS = rl["ParagraphStyle"]
    return {
        "title": PS("t", parent=base["Title"], fontSize=18, spaceAfter=2, textColor=rl["colors"].HexColor("#0f172a")),
        "h1":    PS("h1", parent=base["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=4, textColor=rl["colors"].HexColor("#1e293b")),
        "body":  PS("b", parent=base["BodyText"], fontSize=8.5, leading=11),
        "muted": PS("m", parent=base["BodyText"], fontSize=8, leading=10, textColor=rl["colors"].HexColor("#64748b")),
        "big":   PS("big", parent=base["Title"], fontSize=44, textColor=rl["colors"].HexColor("#4338ca")),
        "ok":    PS("ok", parent=base["BodyText"], fontSize=9, textColor=rl["colors"].HexColor("#047857")),
        "bad":   PS("bad", parent=base["BodyText"], fontSize=9, textColor=rl["colors"].HexColor("#b91c1c")),
    }


def _header(rl, styles, org, subtitle):
    return [
        rl["Paragraph"](f"{org}", styles["title"]),
        rl["Paragraph"](f"ARBITER · {subtitle}", styles["muted"]),
    ]


def _score_table(rl, styles, summaries):
    rows = [["Framework", "Score", "Open", "Passing", "Critical"]]
    for s in summaries:
        rows.append([s["name"], str(s["score"]), str(s["open_count"]),
                     str(s["pass_count"]), str(s["critical_count"])])
    t = rl["Table"](rows, colWidths=[2.6 * rl["inch"], 0.8 * rl["inch"], 0.8 * rl["inch"],
                                     0.9 * rl["inch"], 0.9 * rl["inch"]])
    t.setStyle(rl["TableStyle"]([
        ("BACKGROUND", (0, 0), (-1, 0), rl["colors"].HexColor("#f1f5f9")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOX", (0, 0), (-1, -1), 0.4, rl["colors"].HexColor("#e2e8f0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rl["colors"].HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl["colors"].white, rl["colors"].HexColor("#fafbfc")]),
    ]))
    return t


def _build_pdf(story, rl):
    buf = io.BytesIO()
    doc = rl["SimpleDocTemplate"](buf, pagesize=rl["LETTER"],
                                  leftMargin=0.6 * rl["inch"], rightMargin=0.6 * rl["inch"],
                                  topMargin=0.5 * rl["inch"], bottomMargin=0.5 * rl["inch"])
    doc.build(story)
    return buf.getvalue(), "application/pdf", "pdf"


def executive_compliance(bundle):
    rl = _rl()
    st = _styles(rl)
    summaries = _filter_summaries(bundle["summaries"], bundle.get("params"))
    overall = round(sum(s["score"] for s in summaries) / len(summaries)) if summaries else 0
    has_critical = any(s["critical_count"] for s in summaries)
    story = _header(rl, st, bundle["org_name"], "Executive Compliance Briefing")
    story += [
        rl["Spacer"](1, 0.15 * rl["inch"]),
        rl["Paragraph"]("Overall compliance posture", st["h1"]),
        rl["Paragraph"](f"{overall}", st["big"]),
        rl["Paragraph"](f"Average across {len(summaries)} frameworks · generated {bundle['generated_at']}", st["muted"]),
    ]
    if has_critical:
        story.append(rl["Paragraph"]("<b>ACTION REQUIRED.</b> At least one framework has open CRITICAL conflicts.", st["bad"]))
    else:
        story.append(rl["Paragraph"]("No open critical conflicts across the selected frameworks.", st["ok"]))
    story += [rl["Spacer"](1, 0.12 * rl["inch"]),
              rl["Paragraph"]("Framework scores", st["h1"]),
              _score_table(rl, st, summaries),
              rl["Spacer"](1, 0.18 * rl["inch"]),
              rl["Paragraph"]("Top open risks", st["h1"])]
    top = _top_open_risks(bundle["conflicts"], 5)
    if not top:
        story.append(rl["Paragraph"]("No open risks.", st["body"]))
    else:
        rows = [["ID", "Severity", "Title"]]
        for c in top:
            rows.append([c.get("conflict_id", "—"), c.get("severity", "—"),
                         rl["Paragraph"]((c.get("title") or "")[:95], st["body"])])
        t = rl["Table"](rows, colWidths=[1.2 * rl["inch"], 0.9 * rl["inch"], 4.9 * rl["inch"]])
        t.setStyle(rl["TableStyle"]([
            ("BACKGROUND", (0, 0), (-1, 0), rl["colors"].HexColor("#f1f5f9")),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BOX", (0, 0), (-1, -1), 0.4, rl["colors"].HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl["colors"].white, rl["colors"].HexColor("#fafbfc")]),
        ]))
        story.append(t)
    return _build_pdf(story, rl)


def technical_compliance_pdf(bundle):
    rl = _rl()
    st = _styles(rl)
    summaries = _filter_summaries(bundle["summaries"], bundle.get("params"))
    story = _header(rl, st, bundle["org_name"], "Technical Compliance Report")
    story.append(rl["Paragraph"](f"Per-control posture · generated {bundle['generated_at']}", st["muted"]))
    for s in summaries:
        story += [rl["Spacer"](1, 0.12 * rl["inch"]),
                  rl["Paragraph"](f"{s['name']} — score {s['score']} · {s['pass_count']} passing · {s['open_count']} failing", st["h1"])]
        rows = [["Control", "Name", "Status", "Conflict", "Severity"]]
        for e in s["evals"]:
            linked = e.get("linked") or {}
            rows.append([
                e["ctrl"]["id"],
                rl["Paragraph"](e["ctrl"]["name"], st["body"]),
                e["status"],
                e.get("uc") or "—",
                (e.get("severity") or "—"),
            ])
        t = rl["Table"](rows, colWidths=[1.2 * rl["inch"], 2.6 * rl["inch"], 0.7 * rl["inch"],
                                         1.3 * rl["inch"], 0.9 * rl["inch"]])
        t.setStyle(rl["TableStyle"]([
            ("BACKGROUND", (0, 0), (-1, 0), rl["colors"].HexColor("#f1f5f9")),
            ("FONTSIZE", (0, 0), (-1, -1), 8), ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3), ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOX", (0, 0), (-1, -1), 0.4, rl["colors"].HexColor("#e2e8f0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, rl["colors"].HexColor("#eef2f7")),
        ]))
        story.append(t)
    return _build_pdf(story, rl)


# ──────────────────────────── openpyxl (XLSX) ────────────────────────────
def _xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as e:  # pragma: no cover
        raise RuntimeError(f"XLSX generation requires openpyxl, which is unavailable. ({e})")
    return Workbook, Font, PatternFill


def _xlsx_header(ws, headers, Font, PatternFill):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="F1F5F9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def technical_compliance_xlsx(bundle):
    Workbook, Font, PatternFill = _xlsx()
    summaries = _filter_summaries(bundle["summaries"], bundle.get("params"))
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Scores"
    _xlsx_header(ws1, ["Framework", "Score", "Open", "Passing", "Critical", "High", "Medium"], Font, PatternFill)
    for s in summaries:
        ws1.append([s["name"], s["score"], s["open_count"], s["pass_count"],
                    s["critical_count"], s["high_count"], s["medium_count"]])
    ws2 = wb.create_sheet("Controls")
    _xlsx_header(ws2, ["Framework", "Control", "Name", "Status", "Conflict", "Severity"], Font, PatternFill)
    for s in summaries:
        for e in s["evals"]:
            ws2.append([s["name"], e["ctrl"]["id"], e["ctrl"]["name"], e["status"],
                        e.get("uc") or "", e.get("severity") or ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def conflict_register_xlsx(bundle):
    Workbook, Font, PatternFill = _xlsx()
    conflicts = _filter_conflicts(bundle["conflicts"], bundle.get("params"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Conflicts"
    _xlsx_header(ws, ["ID", "Severity", "Status", "Title", "Domains", "Regulatory", "Detected"], Font, PatternFill)
    for c in conflicts:
        ws.append([
            c.get("conflict_id", ""), c.get("severity", ""), c.get("status", "OPEN"),
            c.get("title", ""), _join(c.get("domains")), _join(c.get("regulatory")),
            c.get("detected_at", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


# ──────────────────────────── CSV / JSON ────────────────────────────
_CONFLICT_FIELDS = ["conflict_id", "severity", "status", "title", "domains",
                    "regulatory", "source_policy", "source_technical", "detected_at"]


def _conflict_register_csv_bytes(conflicts):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_CONFLICT_FIELDS)
    for c in conflicts:
        w.writerow([
            c.get("conflict_id", ""), c.get("severity", ""), c.get("status", "OPEN"),
            c.get("title", ""), _join(c.get("domains")), _join(c.get("regulatory")),
            c.get("source_policy", ""), c.get("source_technical", ""), c.get("detected_at", ""),
        ])
    return buf.getvalue().encode("utf-8")


def conflict_register_csv(bundle):
    conflicts = _filter_conflicts(bundle["conflicts"], bundle.get("params"))
    return _conflict_register_csv_bytes(conflicts), "text/csv", "csv"


def conflict_register_json(bundle):
    conflicts = _filter_conflicts(bundle["conflicts"], bundle.get("params"))
    payload = {"generated_at": bundle["generated_at"], "count": len(conflicts), "conflicts": conflicts}
    return json.dumps(payload, cls=_DecimalEncoder, indent=2).encode("utf-8"), "application/json", "json"


def _audit_csv_bytes(audit):
    buf = io.StringIO()
    fields = ["timestamp", "action_type", "resource", "user", "status", "details"]
    w = csv.writer(buf)
    w.writerow(fields)
    for a in audit:
        details = a.get("details")
        if not isinstance(details, str):
            details = json.dumps(details, cls=_DecimalEncoder)
        w.writerow([a.get("timestamp", ""), a.get("action_type", ""), a.get("resource", ""),
                    a.get("user", ""), a.get("status", ""), details])
    return buf.getvalue().encode("utf-8")


def audit_trail_csv(bundle):
    return _audit_csv_bytes(bundle["audit"]), "text/csv", "csv"


def audit_trail_json(bundle):
    payload = {"generated_at": bundle["generated_at"], "count": len(bundle["audit"]), "events": bundle["audit"]}
    return json.dumps(payload, cls=_DecimalEncoder, indent=2).encode("utf-8"), "application/json", "json"


# ──────────────────────────── ZIP (evidence package) ────────────────────────────
def evidence_package(bundle):
    conflicts = _filter_conflicts(bundle["conflicts"], bundle.get("params"))
    summaries = _filter_summaries(bundle["summaries"], bundle.get("params"))
    scores_json = json.dumps(
        {"generated_at": bundle["generated_at"], "overall": bundle["overall"], "frameworks": summaries},
        cls=_DecimalEncoder, indent=2,
    ).encode("utf-8")
    audit_json = json.dumps(
        {"generated_at": bundle["generated_at"], "events": bundle["audit"]},
        cls=_DecimalEncoder, indent=2,
    ).encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        try:
            pdf_bytes, _, _ = technical_compliance_pdf(bundle)
            zf.writestr("technical_compliance.pdf", pdf_bytes)
        except RuntimeError as e:
            zf.writestr("technical_compliance.SKIPPED.txt", str(e).encode("utf-8"))
        zf.writestr("conflicts.csv", _conflict_register_csv_bytes(conflicts))
        zf.writestr("audit_trail.json", audit_json)
        zf.writestr("scores.json", scores_json)
        zf.writestr("README.txt", (
            f"{bundle['org_name']} — ARBITER evidence package\n"
            f"Generated {bundle['generated_at']}\n\n"
            "Contents:\n"
            "  technical_compliance.pdf  per-framework control posture\n"
            "  conflicts.csv             conflict register\n"
            "  audit_trail.json          audit-log events\n"
            "  scores.json               framework scores snapshot\n"
        ).encode("utf-8"))
    return buf.getvalue(), "application/zip", "zip"


# ──────────────────────────── dispatch ────────────────────────────
GENERATORS = {
    ("executive_compliance", "pdf"): executive_compliance,
    ("technical_compliance", "pdf"): technical_compliance_pdf,
    ("technical_compliance", "xlsx"): technical_compliance_xlsx,
    ("conflict_register", "csv"): conflict_register_csv,
    ("conflict_register", "xlsx"): conflict_register_xlsx,
    ("conflict_register", "json"): conflict_register_json,
    ("audit_trail", "csv"): audit_trail_csv,
    ("audit_trail", "json"): audit_trail_json,
    ("evidence_package", "zip"): evidence_package,
}


def generate(report_id: str, fmt: str, bundle: dict):
    fn = GENERATORS.get((report_id, fmt))
    if not fn:
        raise ValueError(f"No generator for report_type={report_id} format={fmt}")
    return fn(bundle)
